"""
BizReach 統合スクリーニングアプリ（セールスMGR × CS × 若手）
判定基準はサイドバーの「⚙️ 判定基準を編集」から変更できます。
使い方: streamlit run screen_combined.py
"""

import streamlit as st
import pdfplumber
import re
import os
import json
import time
from collections import OrderedDict, Counter
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import anthropic

# ─────────────────────────────────────────────
# 定数
# ─────────────────────────────────────────────
EXCEL_FILENAME  = "スクリーニング管理シート統合.xlsx"
CONFIG_FILENAME = "screening_config.json"

FILL = {
    "A":  PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid"),
    "B+": PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
    "B-": PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
    "C":  PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
}
FONT_BOLD   = Font(bold=True)
FONT_ITALIC = Font(italic=True, color="666666")
JUDGE_EMOJI = {"A": "⭐", "B+": "🔶", "B-": "🔷", "C": "❌"}

# ─────────────────────────────────────────────
# デフォルト設定（構造化データ）
# ─────────────────────────────────────────────
DEFAULT_CONFIG = {
    "mgr": {
        "age_min": 30, "age_max": 50, "jobs_max": 5, "income_max": 1250,
        "ng_patterns": [
            "医療・製薬の旧来型MR・器材営業のみ",
            "大手企業のみでカオス耐性が見えない",
            "業界構造による高年収（銀行・証券・M&A等）",
            "短期離職が多い（1年未満が複数）",
            "デジタル・ITへの接点がゼロ",
        ],
        "high_plus": [
            "ベンチャー・中小で未整備環境から成果を出した経験",
            "仕組み・プロセスの設計・構築経験",
            "エンタープライズ営業 or 代理店営業の経験",
            "IT・SaaS・広告代理店での法人営業",
            "チームマネジメント経験（人数多いほど加点）",
            "提案型営業での継続的な成果",
            "学歴: MARCH・関関同立・旧帝・早慶・上智以上",
        ],
        "medium_plus": [
            "無形商材営業の経験",
            "医療・ヘルスケアへの関心・接点",
            "大手でも新規事業・カオス環境の経験",
            "顧客課題を構造化した提案経験",
        ],
        "judge_a":  "足切りOK + NG無し + 高加点1つ + 中加点1つ",
        "judge_bp": "足切りOK + NG無し + 高加点1つ or 中加点2つ",
        "judge_bm": "足切りOK + NG無し + 中加点1つ",
        "judge_c":  "足切り未クリア or NGパターン複数 or ポテンシャルなし",
        "judge_notes": [
            "迷ったらB+（C方向に引っ張らない）",
            "1つの懸念だけでCにしない",
            "1社でも圧倒的実績があればA検討",
        ],
    },
    "cs": {
        "age_min": 30, "age_max": 50, "jobs_max": 5,
        "income_min": 600, "income_max": 1000,
        "must_skills": [
            "SaaS・IT・広告・人材・Webサービスでの提案型営業またはカスタマーサクセス経験",
            "事業会社・SIerでのプロジェクトマネジメント／PMO経験",
        ],
        "ng_patterns": [
            "旧来型ルート営業・受発注営業のみ",
            "医療MR・医療器材営業のみ（CS経験なし）",
            "デジタル・IT・Webへの接点がゼロ",
            "短期離職が多い（1年未満が複数）",
        ],
        "high_plus": [
            "Webサービスの導入支援経験",
            "オンボーディング設計の経験",
            "データ分析・レポーティング・改善提案の実績",
            "ベンチャーで未整備環境から仕組みを作った経験（最重視）",
            "コンサルティングファームでのプロジェクト推進経験",
            "オペレーション設計・業務プロセス構築の経験",
            "学歴: MARCH・関関同立・旧帝・早慶・上智以上",
        ],
        "medium_plus": [
            "無形商材の提案・コンサル経験",
            "医療・ヘルスケア業界への接点・関心",
            "複数プロジェクトの並行推進経験",
            "アップセル・クロスセルの実績",
            "DX推進・業務改善・業務設計の経験",
            "プロダクト改善・新機能企画への関与",
        ],
        "judge_a":  "足切りOK + MUST充足 + NG無し + 高加点1つ以上 + 中加点1つ以上",
        "judge_bp": "足切りOK + MUSTが弱い + NG無し + 高加点1つ or 中加点2つ",
        "judge_bm": "足切りOK + MUSTが弱い + NG無し + 中加点1つ",
        "judge_c":  "足切り未クリア or NGパターン複数 or MUSTスキル完全不在",
        "judge_notes": [
            "迷ったらB+（C方向に引っ張らない）",
            "懸念点が1つだけならB-ではなくB+にする",
            "SaaS・Web・オンボーディング経験は積極的に評価",
            "医療知識・経験は不問",
        ],
    },
    "junior": {
        "age_min": 25, "age_max": 34, "jobs_max": 3,
        "income_max_29": 700, "income_max_34": 900,
        "ng_patterns": [
            "職歴に具体的な成果・数字の記載がない",
            "ルート営業・受発注営業のみで提案経験なし",
            "短期離職が多い（1年未満が複数）",
            "IT・サービス業への接点がゼロ",
        ],
        "high_plus_atama": [
            "MARCH・関関同立・旧帝・早慶・上智以上の学歴",
            "年次に対して昇進・抜擢が早い",
            "複雑な課題解決・プロジェクトリード経験",
            "論理的思考が職務要約から読み取れる",
        ],
        "high_plus_behavior": [
            "営業成績・行動量が数字で示されている（件数・順位・達成率など）",
            "ベンチャー・スタートアップでの0→1経験",
            "副業・自主学習・社外活動の記述あり",
            "医療・ヘルスケアへの強い関心・動機が明確",
        ],
        "medium_plus": [
            "無形商材・提案型営業の経験",
            "SaaS・IT・広告・人材業界の経験",
            "チームへの貢献・巻き込み経験",
            "医療・ヘルスケアへの接点",
        ],
        "judge_a":  "足切りOK + NG無し + 高加点2つ以上（地頭 or 行動量どちらでも可）",
        "judge_bp": "足切りOK + NG無し + 高加点1つ以上",
        "judge_bm": "足切りOK + NG無し + 中加点1つ以上",
        "judge_c":  "足切り未クリア or NGパターン複数 or ポテンシャル不明",
        "judge_notes": [
            "迷ったらB+（若手は伸びしろで評価）",
            "学歴が高い場合は積極的にプラス評価",
            "行動量・数字の実績がある場合は積極的にプラス評価",
            "医療知識・経験は不問",
        ],
    },
}

# ─────────────────────────────────────────────
# 設定ファイル
# ─────────────────────────────────────────────
def _get_config_path():
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), CONFIG_FILENAME)

def load_config() -> dict:
    path = _get_config_path()
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            cfg = json.load(f)
        # 不足キーを補完
        for pos, defaults in DEFAULT_CONFIG.items():
            if pos not in cfg:
                cfg[pos] = defaults.copy()
            else:
                for k, v in defaults.items():
                    if k not in cfg[pos]:
                        cfg[pos][k] = v
        return cfg
    return {k: v.copy() for k, v in DEFAULT_CONFIG.items()}

def save_config(cfg: dict):
    path = _get_config_path()
    with open(path, "w", encoding="utf-8") as f:
        json.dump(cfg, f, ensure_ascii=False, indent=2)

# ─────────────────────────────────────────────
# プロンプト自動生成
# ─────────────────────────────────────────────
def _items(lst): return "\n".join(f"- {x}" for x in lst if x.strip())

def generate_prompt_mgr(c: dict) -> str:
    return f"""あなたはBizReachの採用スクリーニング専門家です。
株式会社メディカルノートの「プラットフォームセールスMGR」ポジションの候補者を評価します。

【足切り条件】（1つでも外れたらC）
- 年齢: {c['age_min']}〜{c['age_max']}歳
- 転職回数: {c['jobs_max']}社以下
- 現職中であること
- 希望年収: {c['income_max']}万円未満

【NGパターン】（複数該当でC）
{_items(c['ng_patterns'])}

【高加点】（1つで強く評価）
{_items(c['high_plus'])}

【中加点】
{_items(c['medium_plus'])}

【判定基準】
- A: {c['judge_a']}
- B+: {c['judge_bp']}
- B-: {c['judge_bm']}
- C: {c['judge_c']}

【判定ルール】
{_items(c['judge_notes'])}

以下の形式で必ず回答してください：
判定: [A/B+/B-/C]
理由: [2〜3行で根拠を説明]
スカウトポイント: [B+以上の場合、スカウト文に使えるポイントを1行で]"""


def generate_prompt_cs(c: dict) -> str:
    return f"""あなたはBizReachの採用スクリーニング専門家です。
株式会社メディカルノートの「カスタマーサクセス（医療DX）」ポジションの候補者を評価します。

【ポジション概要】
医療機関・製薬企業を顧客とし、サービス導入後の活用促進・オンボーディング・効果測定・アップセルを担う。

【足切り条件】（1つでも外れたらC）
- 年齢: {c['age_min']}〜{c['age_max']}歳
- 転職回数: {c['jobs_max']}社以下
- 現職中であること
- 希望年収: {c['income_min']}万円〜{c['income_max']}万円

【MUSTスキル】（いずれか1つが必要。なければC寄りのB-）
{_items(c['must_skills'])}

【NGパターン】（複数該当でC）
{_items(c['ng_patterns'])}

【高加点】（1つで強く評価）
{_items(c['high_plus'])}

【中加点】
{_items(c['medium_plus'])}

【判定基準】
- A: {c['judge_a']}
- B+: {c['judge_bp']}
- B-: {c['judge_bm']}
- C: {c['judge_c']}

【判定ルール】
{_items(c['judge_notes'])}

以下の形式で必ず回答してください：
判定: [A/B+/B-/C]
理由: [2〜3行で根拠を説明]
スカウトポイント: [B+以上の場合、スカウト文に使えるポイントを1行で]"""


def generate_prompt_junior(c: dict) -> str:
    return f"""あなたはBizReachの採用スクリーニング専門家です。
株式会社メディカルノートの「若手採用（オープン／プラットフォームセールス／事業企画）」ポジションの候補者を評価します。

【ポジション概要】
{c['age_min']}〜{c['age_max']}歳の若手を対象に、ビジネスサイドで活躍できる人材を探します。
経験よりポテンシャル重視。「地頭の高さ」または「素直さ＋行動量」のどちらかがあれば前向きに評価します。

【足切り条件】（1つでも外れたらC）
- 年齢: {c['age_min']}〜{c['age_max']}歳
- 転職回数: {c['jobs_max']}社以下
- 現職中であること
- 希望年収: 25〜29歳は{c['income_max_29']}万以下 / 30〜34歳は{c['income_max_34']}万以下

【NGパターン】（複数該当でC）
{_items(c['ng_patterns'])}

【地頭タイプ 高加点】（1つで強く評価）
{_items(c['high_plus_atama'])}

【素直さ＋行動量タイプ 高加点】（1つで強く評価）
{_items(c['high_plus_behavior'])}

【中加点】
{_items(c['medium_plus'])}

【判定基準】
- A: {c['judge_a']}
- B+: {c['judge_bp']}
- B-: {c['judge_bm']}
- C: {c['judge_c']}

【判定ルール】
{_items(c['judge_notes'])}

【推奨求人の選定基準】
- プラットフォームセールス: 営業経験が中心の候補者
- 事業企画: 学歴が高い（MARCH以上）または企画・事業開発・PMの経験が中心の候補者
- オープンポジション: 上記どちらにも当てはまらない候補者

以下の形式で必ず回答してください：
判定: [A/B+/B-/C]
理由: [2〜3行で根拠を説明]
スカウトポイント: [B+以上の場合、スカウト文に使えるポイントを1行で]
推奨求人: [プラットフォームセールス/事業企画/オープンポジション]
推奨理由: [1行で理由]"""


# ─────────────────────────────────────────────
# PDF 処理
# ─────────────────────────────────────────────
def extract_candidates_from_pdf(uploaded_file) -> OrderedDict:
    full_text = ""
    with pdfplumber.open(uploaded_file) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t:
                full_text += t + "\n"
    parts = re.split(r"(?=^BU\d{7}\n)", full_text, flags=re.MULTILINE)
    candidates = OrderedDict()
    for part in parts:
        m = re.match(r"^(BU\d{7})\n", part)
        if m:
            bu_id = m.group(1)
            candidates[bu_id] = candidates.get(bu_id, "") + "\n" + part
    return candidates


def extract_profile(text: str) -> dict:
    age_m  = re.search(r"(男性|女性)\s*/\s*(\d+)歳", text)
    gender = age_m.group(1) if age_m else "不明"
    age    = int(age_m.group(2)) if age_m else 0
    inc_m  = re.search(r"直近の年収\s+([\d,万〜円]+)", text)
    income = inc_m.group(1) if inc_m else "不明"
    edu_m  = re.search(r"学歴\s+•\s+(.+?)(?:\s+\d{4}年|\s*\n)", text)
    edu    = edu_m.group(1).strip() if edu_m else "不明"
    blk    = re.search(r"所属企業一覧(.*?)直近の年収", text, re.DOTALL)
    companies, titles = [], []
    if blk:
        for line in blk.group(1).strip().split("\n"):
            if "•" in line:
                m = re.search(r"•\s+(.+?)(?:\s{2,}|\d{4}年)", line)
                if m:
                    raw = m.group(1).strip()
                    sp  = raw.split(" / ", 1)
                    companies.append(sp[0].strip())
                    titles.append(sp[1].strip() if len(sp) > 1 else "")
    mgt_m = re.search(r"マネジメント経験\s+(.+?)(?:\n|$)", text)
    mgt   = mgt_m.group(1).strip() if mgt_m else "記載なし"
    exp_m = re.search(r"経験職種(.*?)経験業種", text, re.DOTALL)
    exp_jobs = []
    if exp_m:
        for line in exp_m.group(1).strip().split("\n"):
            if "•" in line:
                j = re.search(r"•\s+(.+?)(?:\s+経験年数|\s*$)", line)
                if j: exp_jobs.append(j.group(1).strip())
    ind_m = re.search(r"経験業種(.*?)マネジメント経験", text, re.DOTALL)
    industries = []
    if ind_m:
        for line in ind_m.group(1).strip().split("\n"):
            if "•" in line:
                i = re.search(r"•\s+(.+?)(?:\s+経験年数|\s*$)", line)
                if i: industries.append(i.group(1).strip())
    sum_m = re.search(r"職務要約\n(.*?)(?:職務経歴書の取り扱い|職務経歴\n|$)", text, re.DOTALL)
    job_summary = sum_m.group(1).strip()[:600] if sum_m else ""
    return {
        "age": age, "gender": gender, "income": income, "edu": edu,
        "current_company": companies[0] if companies else "不明",
        "current_title":   titles[0]   if titles   else "",
        "num_jobs": len(companies),
        "companies": companies, "titles": titles,
        "mgt": mgt, "job_summary": job_summary,
        "exp_jobs": exp_jobs, "industries": industries,
    }


def parse_income_value(income_str: str):
    if "〜" in income_str:
        return None
    m = re.search(r"([\d,]+)万", income_str)
    if m:
        return int(m.group(1).replace(",", ""))
    return None


# ─────────────────────────────────────────────
# 自動足切り
# ─────────────────────────────────────────────
def auto_cut_mgr(profile: dict, cfg: dict):
    c = cfg["mgr"]
    age, n_jobs = profile["age"], profile["num_jobs"]
    if age > 0 and (age < c["age_min"] or age > c["age_max"]):
        return "C", f"年齢{age}歳（{c['age_min']}〜{c['age_max']}歳の範囲外）"
    if n_jobs == 1:
        return "C", "在籍1社のみ"
    if n_jobs > c["jobs_max"]:
        return "C", f"転職{n_jobs}社（上限{c['jobs_max']}社超過）"
    val = parse_income_value(profile["income"])
    if val and val >= c["income_max"]:
        return "C", f"希望年収 {profile['income']}（{c['income_max']}万以上 NG）"
    return None, ""


def auto_cut_cs(profile: dict, cfg: dict):
    c = cfg["cs"]
    age, n_jobs = profile["age"], profile["num_jobs"]
    if age > 0 and (age < c["age_min"] or age > c["age_max"]):
        return "C", f"年齢{age}歳（{c['age_min']}〜{c['age_max']}歳の範囲外）"
    if n_jobs == 1:
        return "C", "在籍1社のみ"
    if n_jobs > c["jobs_max"]:
        return "C", f"転職{n_jobs}社（上限{c['jobs_max']}社超過）"
    val = parse_income_value(profile["income"])
    if val:
        if val < c["income_min"]:
            return "C", f"希望年収 {profile['income']}（{c['income_min']}万未満 NG）"
        if val > c["income_max"]:
            return "C", f"希望年収 {profile['income']}（{c['income_max']}万超 NG）"
    return None, ""


def auto_cut_junior(profile: dict, cfg: dict):
    c = cfg["junior"]
    age, n_jobs = profile["age"], profile["num_jobs"]
    if age == 0 or age < c["age_min"] or age > c["age_max"]:
        return "C", f"年齢{age}歳（{c['age_min']}〜{c['age_max']}歳の範囲外）"
    if n_jobs > c["jobs_max"]:
        return "C", f"転職{n_jobs}社（上限{c['jobs_max']}社超過）"
    val = parse_income_value(profile["income"])
    if val:
        if age <= 29 and val > c["income_max_29"]:
            return "C", f"希望年収 {profile['income']}（25〜29歳は{c['income_max_29']}万超 NG）"
        if age >= 30 and val > c["income_max_34"]:
            return "C", f"希望年収 {profile['income']}（30〜34歳は{c['income_max_34']}万超 NG）"
    return None, ""


# ─────────────────────────────────────────────
# Claude API 判定
# ─────────────────────────────────────────────
def ai_judge(api_key: str, bu_id: str, profile: dict, system_prompt: str,
             _status_placeholder=None) -> dict:
    client = anthropic.Anthropic(api_key=api_key)
    candidate_info = f"""
候補者ID: {bu_id}
年齢・性別: {profile['age']}歳 {profile['gender']}
現職: {profile['current_company']} / {profile['current_title']}
年収: {profile['income']}
学歴: {profile['edu']}
転職回数: {profile['num_jobs']}社
マネジメント: {profile['mgt']}
経験職種: {', '.join(profile['exp_jobs'][:5])}
経験業種: {', '.join(profile['industries'][:5])}

【職歴】
{chr(10).join(f'・{c} / {t}' for c, t in zip(profile['companies'], profile['titles']))}

【職務要約】
{profile['job_summary']}
"""
    # レート制限対策: 最大3回リトライ（待機時間: 65秒 → 130秒 → 195秒）
    max_retries = 3
    for attempt in range(max_retries + 1):
        try:
            response = client.messages.create(
                model="claude-haiku-4-5-20251001",
                max_tokens=300,
                system=system_prompt,
                messages=[{"role": "user", "content": candidate_info}]
            )
            result_text = response.content[0].text
            judge_m  = re.search(r"判定:\s*([AB][+\-]?|C)", result_text)
            reason_m = re.search(r"理由:\s*(.+?)(?:スカウトポイント:|$)", result_text, re.DOTALL)
            scout_m  = re.search(r"スカウトポイント:\s*(.+?)$", result_text, re.DOTALL)
            judge  = judge_m.group(1).strip()  if judge_m  else "B-"
            reason = reason_m.group(1).strip() if reason_m else result_text
            scout  = scout_m.group(1).strip()  if scout_m  else ""
            if judge not in ["A", "B+", "B-", "C"]:
                judge = "B-"
            return {"judge": judge, "reason": reason, "scout": scout, "raw": result_text}

        except Exception as e:
            err_str = str(e)
            is_rate_limit = (
                "rate_limit" in err_str.lower() or
                "429" in err_str or
                "too_many_requests" in err_str.lower() or
                "overloaded" in err_str.lower()
            )
            if is_rate_limit and attempt < max_retries:
                wait_sec = 65 * (attempt + 1)  # 65秒 → 130秒 → 195秒
                for remaining in range(wait_sec, 0, -1):
                    if _status_placeholder:
                        _status_placeholder.warning(
                            f"⏳ レート制限のため待機中… {remaining}秒後にリトライします "
                            f"（{attempt+1}/{max_retries}回目）　`{bu_id}`"
                        )
                    time.sleep(1)
            else:
                # リトライ上限超過 or レート制限以外のエラー
                return {"judge": "ERROR", "reason": f"APIエラー（{attempt+1}回試行）: {err_str[:120]}", "scout": "", "raw": ""}


def ai_judge_junior(api_key: str, bu_id: str, profile: dict, system_prompt: str,
                    _status_placeholder=None) -> dict:
    """若手専用: 通常の判定に加え推奨求人も取得する"""
    result = ai_judge(api_key, bu_id, profile, system_prompt, _status_placeholder=_status_placeholder)
    raw = result.get("raw", "")
    job_m    = re.search(r"推奨求人:\s*(.+?)(?:\n|$)", raw)
    reason_m = re.search(r"推奨理由:\s*(.+?)(?:\n|$)", raw)
    scout_job    = job_m.group(1).strip()    if job_m    else "オープンポジション"
    scout_reason = reason_m.group(1).strip() if reason_m else ""
    # 正規化
    if "セールス" in scout_job or "営業" in scout_job:
        scout_job = "プラットフォームセールス"
    elif "企画" in scout_job or "事業" in scout_job:
        scout_job = "事業企画"
    else:
        scout_job = "オープンポジション"
    result["scout_job"]    = scout_job
    result["scout_reason"] = scout_reason
    return result


# ─────────────────────────────────────────────
# Excel 操作
# ─────────────────────────────────────────────
HEADERS = [
    "No.", "BU ID", "年齢", "性別", "会社名", "役職", "年収", "学歴", "転職回数",
    "MGR判定", "MGR理由", "MGRスカウトポイント",
    "CS判定",  "CS理由",  "CSスカウトポイント",
    "若手判定", "若手理由", "若手スカウトポイント", "若手推奨求人",
    "備考"
]
COL_WIDTHS = [5, 14, 5, 5, 28, 22, 16, 26, 8, 8, 40, 36, 8, 40, 36, 8, 40, 36, 22, 18]


def _get_excel_path():
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), EXCEL_FILENAME)


def load_existing_ids():
    path = _get_excel_path()
    if not os.path.exists(path):
        return set(), 0
    wb = load_workbook(path)
    ws = wb.active
    ids, max_no = set(), 0
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[0] and str(row[0]).isdigit():
            max_no = max(max_no, int(row[0]))
        if row[1] and str(row[1]).startswith("BU"):
            ids.add(str(row[1]))
    return ids, max_no


def _create_excel(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "統合スクリーニング"
    ws["A1"] = "BizReach 統合スクリーニング管理シート｜セールスMGR × カスタマーサクセス × 若手"
    ws["A2"] = "集計: (自動更新)"
    ws["A3"] = "※ MGR判定：セールスMGR　CS判定：カスタマーサクセス（医療DX）　若手判定：若手採用"
    ws.merge_cells("A1:T1")
    ws.merge_cells("A2:T2")
    ws.merge_cells("A3:T3")
    hf = PatternFill(start_color="1C2833", end_color="1C2833", fill_type="solid")
    for col, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.fill = hf
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[4].height = 24
    ws.freeze_panes = "A5"
    wb.save(path)


def append_to_excel(results: list) -> tuple:
    path = _get_excel_path()
    if not os.path.exists(path):
        _create_excel(path)
    wb = load_workbook(path)
    ws = wb.active
    existing_ids, max_no = set(), 0
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[0] and str(row[0]).isdigit():
            max_no = max(max_no, int(row[0]))
        if row[1] and str(row[1]).startswith("BU"):
            existing_ids.add(str(row[1]))
    added = 0
    for item in results:
        if item["bu_id"] in existing_ids:
            continue
        max_no += 1
        target_row = ws.max_row + 1
        row_data = [
            max_no, item["bu_id"],
            item["age"], item["gender"],
            item["company"], item["title"],
            item["income"], item["edu"], item["jobs"],
            item["mgr"]["judge"],    item["mgr"]["reason"],    item["mgr"]["scout"],
            item["cs"]["judge"],     item["cs"]["reason"],     item["cs"]["scout"],
            item["junior"]["judge"], item["junior"]["reason"], item["junior"]["scout"],
            item["junior"].get("scout_job", ""),
            item.get("note", ""),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=target_row, column=col, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for col, key in [(10, "mgr"), (13, "cs"), (16, "junior")]:
            jc = ws.cell(row=target_row, column=col)
            j  = item[key]["judge"]
            if j in FILL: jc.fill = FILL[j]
            jc.font = FONT_BOLD if j in ("A","B+") else (FONT_ITALIC if j == "B-" else Font())
        added += 1
    mgr_cnt = {"A":0,"B+":0,"B-":0,"C":0}
    cs_cnt  = {"A":0,"B+":0,"B-":0,"C":0}
    jr_cnt  = {"A":0,"B+":0,"B-":0,"C":0}
    total   = 0
    for row in ws.iter_rows(min_row=5, values_only=True):
        if row[0] and str(row[0]).isdigit():
            total += 1
            if row[9]  in mgr_cnt: mgr_cnt[row[9]]  += 1
            if row[12] in cs_cnt:  cs_cnt[row[12]]  += 1
            if row[15] in jr_cnt:  jr_cnt[row[15]]  += 1
    ws["A2"] = (
        f"累計{total}名  ｜  "
        f"【MGR】A:{mgr_cnt['A']} B+:{mgr_cnt['B+']} B-:{mgr_cnt['B-']} C:{mgr_cnt['C']}  ｜  "
        f"【CS】A:{cs_cnt['A']} B+:{cs_cnt['B+']} B-:{cs_cnt['B-']} C:{cs_cnt['C']}  ｜  "
        f"【若手】A:{jr_cnt['A']} B+:{jr_cnt['B+']} B-:{jr_cnt['B-']} C:{jr_cnt['C']}"
    )
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    base, ext = os.path.splitext(path)
    new_path  = f"{base}_{timestamp}{ext}"
    wb.save(path)
    wb.save(new_path)
    return added, total, new_path


# ─────────────────────────────────────────────
# 判定基準設定ページ
# ─────────────────────────────────────────────
def _list_editor(label: str, items: list, key: str, help_text: str = "") -> list:
    """1行1項目のテキストエリアで項目リストを編集する。追加・削除は行の追加・削除で行う。"""
    text = "\n".join(items)
    edited = st.text_area(
        label,
        value=text,
        height=max(80, min(300, len(items) * 38 + 40)),
        key=key,
        help=help_text or "1行に1項目。行を追加すると項目が増え、行を削除すると項目が消えます。",
    )
    return [line.strip() for line in edited.split("\n") if line.strip()]


def page_settings():
    st.title("⚙️ 判定基準を編集")
    st.caption("各項目を編集して「保存」を押すと次回判定から反映されます。")

    cfg = load_config()
    tab_mgr, tab_cs, tab_jr = st.tabs(["🏢 セールスMGR", "🏥 カスタマーサクセス", "🌱 若手"])

    # ── MGR ──────────────────────────────────
    with tab_mgr:
        c = cfg["mgr"]
        st.subheader("📋 足切り条件（自動でCになる）")
        col1, col2, col3, col4 = st.columns(4)
        c["age_min"]    = col1.number_input("年齢 下限", value=int(c["age_min"]),    min_value=18, max_value=65, key="mgr_age_min")
        c["age_max"]    = col2.number_input("年齢 上限", value=int(c["age_max"]),    min_value=18, max_value=65, key="mgr_age_max")
        c["jobs_max"]   = col3.number_input("転職回数 上限", value=int(c["jobs_max"]), min_value=1,  max_value=10, key="mgr_jobs_max")
        c["income_max"] = col4.number_input("希望年収 上限（万）", value=int(c["income_max"]), min_value=300, max_value=3000, step=50, key="mgr_income_max")

        st.divider()
        st.subheader("❌ NGパターン")
        st.caption("複数該当でC判定。1行1項目で記入。")
        c["ng_patterns"] = _list_editor("NGパターン", c["ng_patterns"], "mgr_ng")

        st.divider()
        st.subheader("⬆️ 高加点（1つで強く評価）")
        c["high_plus"] = _list_editor("高加点", c["high_plus"], "mgr_hp")

        st.divider()
        st.subheader("➕ 中加点")
        c["medium_plus"] = _list_editor("中加点", c["medium_plus"], "mgr_mp")

        st.divider()
        st.subheader("🏅 判定基準")
        c["judge_a"]  = st.text_input("⭐ A の条件",  value=c["judge_a"],  key="mgr_ja")
        c["judge_bp"] = st.text_input("🔶 B+ の条件", value=c["judge_bp"], key="mgr_jbp")
        c["judge_bm"] = st.text_input("🔷 B- の条件", value=c["judge_bm"], key="mgr_jbm")
        c["judge_c"]  = st.text_input("❌ C の条件",  value=c["judge_c"],  key="mgr_jc")

        st.divider()
        st.subheader("📌 判定ルール")
        st.caption("AIへの補足指示。迷ったときの方針など。")
        c["judge_notes"] = _list_editor("判定ルール", c["judge_notes"], "mgr_jn")

        cfg["mgr"] = c

    # ── CS ──────────────────────────────────
    with tab_cs:
        c = cfg["cs"]
        st.subheader("📋 足切り条件（自動でCになる）")
        col1, col2, col3, col4, col5 = st.columns(5)
        c["age_min"]    = col1.number_input("年齢 下限", value=int(c["age_min"]),    min_value=18, max_value=65, key="cs_age_min")
        c["age_max"]    = col2.number_input("年齢 上限", value=int(c["age_max"]),    min_value=18, max_value=65, key="cs_age_max")
        c["jobs_max"]   = col3.number_input("転職回数 上限", value=int(c["jobs_max"]), min_value=1,  max_value=10, key="cs_jobs_max")
        c["income_min"] = col4.number_input("希望年収 下限（万）", value=int(c["income_min"]), min_value=100, max_value=2000, step=50, key="cs_income_min")
        c["income_max"] = col5.number_input("希望年収 上限（万）", value=int(c["income_max"]), min_value=100, max_value=3000, step=50, key="cs_income_max")

        st.divider()
        st.subheader("✅ MUSTスキル")
        st.caption("いずれか1つが必要。なければB-寄りに判定。")
        c["must_skills"] = _list_editor("MUSTスキル", c["must_skills"], "cs_must")

        st.divider()
        st.subheader("❌ NGパターン")
        c["ng_patterns"] = _list_editor("NGパターン", c["ng_patterns"], "cs_ng")

        st.divider()
        st.subheader("⬆️ 高加点（1つで強く評価）")
        c["high_plus"] = _list_editor("高加点", c["high_plus"], "cs_hp")

        st.divider()
        st.subheader("➕ 中加点")
        c["medium_plus"] = _list_editor("中加点", c["medium_plus"], "cs_mp")

        st.divider()
        st.subheader("🏅 判定基準")
        c["judge_a"]  = st.text_input("⭐ A の条件",  value=c["judge_a"],  key="cs_ja")
        c["judge_bp"] = st.text_input("🔶 B+ の条件", value=c["judge_bp"], key="cs_jbp")
        c["judge_bm"] = st.text_input("🔷 B- の条件", value=c["judge_bm"], key="cs_jbm")
        c["judge_c"]  = st.text_input("❌ C の条件",  value=c["judge_c"],  key="cs_jc")

        st.divider()
        st.subheader("📌 判定ルール")
        c["judge_notes"] = _list_editor("判定ルール", c["judge_notes"], "cs_jn")

        cfg["cs"] = c

    # ── 若手 ──────────────────────────────────
    with tab_jr:
        c = cfg["junior"]
        st.subheader("📋 足切り条件（自動でCになる）")
        col1, col2, col3, col4, col5 = st.columns(5)
        c["age_min"]       = col1.number_input("年齢 下限", value=int(c["age_min"]),       min_value=18, max_value=65, key="jr_age_min")
        c["age_max"]       = col2.number_input("年齢 上限", value=int(c["age_max"]),       min_value=18, max_value=65, key="jr_age_max")
        c["jobs_max"]      = col3.number_input("転職回数 上限", value=int(c["jobs_max"]),   min_value=1,  max_value=10, key="jr_jobs_max")
        c["income_max_29"] = col4.number_input("29歳以下 年収上限（万）", value=int(c["income_max_29"]), min_value=100, max_value=2000, step=50, key="jr_inc29")
        c["income_max_34"] = col5.number_input("30〜34歳 年収上限（万）", value=int(c["income_max_34"]), min_value=100, max_value=2000, step=50, key="jr_inc34")

        st.divider()
        st.subheader("❌ NGパターン")
        c["ng_patterns"] = _list_editor("NGパターン", c["ng_patterns"], "jr_ng")

        st.divider()
        st.subheader("⬆️ 高加点：地頭タイプ")
        c["high_plus_atama"] = _list_editor("地頭タイプ", c["high_plus_atama"], "jr_hp_a")

        st.divider()
        st.subheader("⬆️ 高加点：素直さ＋行動量タイプ")
        c["high_plus_behavior"] = _list_editor("行動量タイプ", c["high_plus_behavior"], "jr_hp_b")

        st.divider()
        st.subheader("➕ 中加点")
        c["medium_plus"] = _list_editor("中加点", c["medium_plus"], "jr_mp")

        st.divider()
        st.subheader("🏅 判定基準")
        c["judge_a"]  = st.text_input("⭐ A の条件",  value=c["judge_a"],  key="jr_ja")
        c["judge_bp"] = st.text_input("🔶 B+ の条件", value=c["judge_bp"], key="jr_jbp")
        c["judge_bm"] = st.text_input("🔷 B- の条件", value=c["judge_bm"], key="jr_jbm")
        c["judge_c"]  = st.text_input("❌ C の条件",  value=c["judge_c"],  key="jr_jc")

        st.divider()
        st.subheader("📌 判定ルール")
        c["judge_notes"] = _list_editor("判定ルール", c["judge_notes"], "jr_jn")

        cfg["junior"] = c

    st.divider()
    if st.button("💾 設定を保存", type="primary", use_container_width=True):
        save_config(cfg)
        st.success("✅ 保存しました。次回の判定から反映されます。")
        st.caption("プロンプトはこの設定から自動生成されます。")


# ─────────────────────────────────────────────
# スクリーニングページ
# ─────────────────────────────────────────────
def page_screening():
    st.title("🔍 統合スクリーニング｜セールスMGR × CS × 若手")
    st.caption("1つのPDFで3ポジションを同時判定します")

    cfg = load_config()

    for k, v in [("results", None), ("saved", False), ("pdf_name", "")]:
        if k not in st.session_state:
            st.session_state[k] = v

    with st.sidebar:
        st.header("⚙️ 設定")
        api_key = st.text_input("🔑 Anthropic API キー", type="password", placeholder="sk-ant-...")
        st.divider()
        excel_path = _get_excel_path()
        if os.path.exists(excel_path):
            existing_ids, max_no = load_existing_ids()
            st.success("✅ Excel 読み込み済み")
            st.info(f"累計 **{max_no}** 名登録済み")
        else:
            existing_ids, max_no = set(), 0
            st.warning("Excel が見つかりません（初回は自動作成）")
        st.divider()
        c_mgr, c_cs, c_jr = cfg["mgr"], cfg["cs"], cfg["junior"]
        st.markdown(f"""**📋 現在の足切り条件**
**MGR**: {c_mgr['age_min']}〜{c_mgr['age_max']}歳 / {c_mgr['jobs_max']}社以下 / {c_mgr['income_max']}万未満
**CS**: {c_cs['age_min']}〜{c_cs['age_max']}歳 / {c_cs['jobs_max']}社以下 / {c_cs['income_min']}〜{c_cs['income_max']}万
**若手**: {c_jr['age_min']}〜{c_jr['age_max']}歳 / {c_jr['jobs_max']}社以下 / 29歳以下:{c_jr['income_max_29']}万 / 30〜34歳:{c_jr['income_max_34']}万""")
        st.divider()
        uploaded_files = st.file_uploader("📄 PDF をアップロード（複数可）", type=["pdf"], accept_multiple_files=True)
        pdf_key = ",".join(sorted(f.name for f in uploaded_files)) if uploaded_files else ""
        if uploaded_files and api_key and pdf_key != st.session_state.pdf_name:
            if st.button("🚀 3ポジション同時判定スタート", type="primary", use_container_width=True):
                st.session_state.pdf_name = pdf_key
                st.session_state.saved    = False
                st.session_state.results  = None
                st.rerun()
        if not api_key:
            st.warning("APIキーを入力してください")

    if not uploaded_files:
        st.info("← サイドバーからAPIキーを入力し、PDFをアップロードしてください")
        return
    if not api_key:
        st.warning("← APIキーを入力してください")
        return

    # ── 判定実行 ──────────────────────────────────
    if st.session_state.results is None and pdf_key == st.session_state.pdf_name:
        with st.spinner(f"📄 PDF {len(uploaded_files)}枚を読み込み中..."):
            candidates = OrderedDict()
            for f in uploaded_files:
                for bu_id, text in extract_candidates_from_pdf(f).items():
                    if bu_id not in candidates:
                        candidates[bu_id] = text
        new_candidates = {k: v for k, v in candidates.items() if k not in existing_ids}
        dup_count = len(candidates) - len(new_candidates)
        total_n   = len(new_candidates)
        st.info(f"PDF {len(uploaded_files)}枚　抽出: {len(candidates)}名 ／ 重複除外: {dup_count}名 ／ 判定対象: **{total_n}名**")

        prompt_mgr    = generate_prompt_mgr(cfg["mgr"])
        prompt_cs     = generate_prompt_cs(cfg["cs"])
        prompt_junior = generate_prompt_junior(cfg["junior"])

        results      = []
        progress_bar = st.progress(0)
        status_text  = st.empty()

        for i, (bu_id, text) in enumerate(new_candidates.items()):
            profile = extract_profile(text)
            progress_bar.progress((i + 1) / total_n)
            status_text.markdown(f"判定中… **{i+1}/{total_n}**　`{bu_id}`")

            def judge_pos(auto_cut_fn, prompt):
                j, note = auto_cut_fn(profile, cfg)
                if j == "C":
                    return {"judge": "C", "reason": note, "scout": ""}
                res = ai_judge(api_key, bu_id, profile, prompt, _status_placeholder=status_text)
                time.sleep(0.5)
                if res["judge"] == "ERROR":
                    res["judge"] = "B-"
                return res

            # 若手は推奨求人も判定する専用関数を使う
            def judge_junior():
                j, note = auto_cut_junior(profile, cfg)
                if j == "C":
                    return {"judge": "C", "reason": note, "scout": "", "scout_job": "—", "scout_reason": ""}
                res = ai_judge_junior(api_key, bu_id, profile, prompt_junior, _status_placeholder=status_text)
                time.sleep(0.5)
                if res["judge"] == "ERROR":
                    res["judge"] = "B-"
                return res

            results.append({
                "bu_id":   bu_id,
                "age":     profile["age"],  "gender":  profile["gender"],
                "company": profile["current_company"],
                "title":   profile["current_title"],
                "income":  profile["income"], "edu": profile["edu"],
                "jobs":    str(profile["num_jobs"]) + "社",
                "mgr":     judge_pos(auto_cut_mgr, prompt_mgr),
                "cs":      judge_pos(auto_cut_cs,  prompt_cs),
                "junior":  judge_junior(),
                "profile": profile,
                "note":    "",
            })

        progress_bar.empty()
        status_text.empty()
        st.session_state.results = results
        st.success(f"✅ 判定完了！{total_n}名 × 3ポジション")
        st.rerun()

    # ── 結果表示 ──────────────────────────────────
    results = st.session_state.results
    if results is None:
        return

    mgr_cnt = Counter(r["mgr"]["judge"]    for r in results)
    cs_cnt  = Counter(r["cs"]["judge"]     for r in results)
    jr_cnt  = Counter(r["junior"]["judge"] for r in results)

    st.subheader("📊 判定サマリー")
    col1, col2, col3 = st.columns(3)
    for col, label, cnt in [
        (col1, "🏢 セールスMGR",       mgr_cnt),
        (col2, "🏥 カスタマーサクセス", cs_cnt),
        (col3, "🌱 若手（25〜34歳）",  jr_cnt),
    ]:
        with col:
            st.markdown(f"**{label}**")
            c1, c2, c3, c4 = st.columns(4)
            c1.metric("⭐ A",  cnt.get("A",  0))
            c2.metric("🔶 B+", cnt.get("B+", 0))
            c3.metric("🔷 B-", cnt.get("B-", 0))
            c4.metric("❌ C",  cnt.get("C",  0))

    st.divider()

    tab_all, tab_mgr_ab, tab_cs_ab, tab_jr_ab, tab_multi_ab = st.tabs([
        f"👥 全員 ({len(results)}名)",
        f"🏢 MGR A/B+ ({sum(1 for r in results if r['mgr']['judge'] in ['A','B+'])}名)",
        f"🏥 CS A/B+ ({sum(1 for r in results if r['cs']['judge'] in ['A','B+'])}名)",
        f"🌱 若手 A/B+ ({sum(1 for r in results if r['junior']['judge'] in ['A','B+'])}名)",
        f"⭐ 複数A/B+ ({sum(1 for r in results if sum(1 for k in ['mgr','cs','junior'] if r[k]['judge'] in ['A','B+']) >= 2)}名)",
    ])

    def show_cards(tab, filtered, tab_key):
        with tab:
            if not filtered:
                st.caption("該当なし")
                return
            for r in filtered:
                p    = r.get("profile", {})
                j_mgr, j_cs, j_jr = r["mgr"]["judge"], r["cs"]["judge"], r["junior"]["judge"]
                with st.expander(
                    f"**{r['bu_id']}**　{r['age']}歳{r['gender']}　{r['company'][:18]}　｜　"
                    f"MGR:{JUDGE_EMOJI[j_mgr]}{j_mgr}　CS:{JUDGE_EMOJI[j_cs]}{j_cs}　若手:{JUDGE_EMOJI[j_jr]}{j_jr}"
                ):
                    c1, c2, c3 = st.columns(3)
                    for col, pos_label, j, key in [
                        (c1, "🏢 MGR",  j_mgr, "mgr"),
                        (c2, "🏥 CS",   j_cs,  "cs"),
                        (c3, "🌱 若手", j_jr,  "junior"),
                    ]:
                        with col:
                            st.markdown(f"### {pos_label}　{JUDGE_EMOJI[j]} **{j}**")
                            st.markdown(f"**📋 理由:**\n{r[key]['reason']}")
                            if r[key]["scout"]:
                                st.markdown(f"**✉️** {r[key]['scout']}")
                            if key == "junior" and r[key].get("scout_job") and r[key]["scout_job"] != "—":
                                st.info(f"📌 推奨求人：**{r[key]['scout_job']}**　{r[key].get('scout_reason','')}")
                            new_j = st.selectbox(f"{pos_label}判定を変更",
                                ["（変更しない）","A","B+","B-","C"],
                                key=f"{key}_{tab_key}_{r['bu_id']}")
                            if new_j != "（変更しない）":
                                r[key]["judge"] = new_j
                    st.markdown("---")
                    ci1, ci2 = st.columns([3, 2])
                    with ci1:
                        st.markdown(f"**役職:** {r['title']}　**学歴:** {r['edu']}")
                        if p.get("companies"):
                            st.markdown("**職歴:** " + "　→　".join(
                                f"{c}（{t}）" for c, t in zip(p["companies"][:3], p["titles"][:3])
                            ))
                    with ci2:
                        st.markdown(f"**年収:** {r['income']}　**転職:** {r['jobs']}")
                        st.markdown(f"**MGT:** {p.get('mgt','')}")
                        if p.get("exp_jobs"):
                            st.markdown(f"**経験職種:** {', '.join(p['exp_jobs'][:3])}")

    show_cards(tab_all,    results, "all")
    show_cards(tab_mgr_ab, [r for r in results if r["mgr"]["judge"]    in ["A","B+"]], "mgr")
    show_cards(tab_cs_ab,  [r for r in results if r["cs"]["judge"]     in ["A","B+"]], "cs")
    show_cards(tab_jr_ab,  [r for r in results if r["junior"]["judge"] in ["A","B+"]], "jr")
    show_cards(tab_multi_ab,
        [r for r in results if sum(1 for k in ["mgr","cs","junior"] if r[k]["judge"] in ["A","B+"]) >= 2],
        "multi")

    st.divider()
    col_save, col_reset = st.columns([3, 1])
    with col_save:
        if not st.session_state.saved:
            if st.button("💾　Excelに追記する", type="primary", use_container_width=True):
                added, total_count, new_path = append_to_excel(results)
                st.session_state.saved = True
                st.balloons()
                st.success(f"✅ {added}名を追記しました（累計 {total_count}名）\n📁 {os.path.basename(new_path)}")
        else:
            st.success("✅ 保存済みです")
    with col_reset:
        if st.button("🔄 次のPDF", use_container_width=True):
            st.session_state.results  = None
            st.session_state.saved    = False
            st.session_state.pdf_name = ""
            st.rerun()


# ─────────────────────────────────────────────
# メイン
# ─────────────────────────────────────────────
def main():
    st.set_page_config(page_title="統合スクリーニング", layout="wide", page_icon="🔍")
    page = st.sidebar.radio(
        "ページ",
        ["🔍 スクリーニング", "⚙️ 判定基準を編集"],
        label_visibility="collapsed"
    )
    if page == "🔍 スクリーニング":
        page_screening()
    else:
        page_settings()


if __name__ == "__main__":
    main()
